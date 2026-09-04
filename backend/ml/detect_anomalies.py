import pandas as pd
import numpy as np
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import LabelEncoder
import openpyxl
import json
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'berkeley-data.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'anomaly_results.json')

TYPE_MAP = {
    'REDD+': 'REDD+', 'Jurisdictional REDD+': 'REDD+',
    'Improved Forest Management': 'IFM',
    'Afforestation/Reforestation': 'ARR',
    'Biochar': 'Biochar', 'Enhanced Rock Weathering': 'ERW',
    'Direct Air Capture': 'DAC'
}

def load_data():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb['PROJECTS']
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    headers = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    col = {h: i for i, h in enumerate(headers) if h}
    credits_col = next((i for h, i in col.items() if str(h).replace('\n',' ').strip().startswith('Total Credits')), None)

    data = []
    for row in rows:
        if not row[col['Project Name']]:
            continue
        t = TYPE_MAP.get(row[col['Type']])
        if not t:
            continue
        vintage = row[col['First Year of Project (Vintage)']]
        credits = row[credits_col] if credits_col is not None else None
        registry = row[col['Voluntary Registry']]
        if not vintage or not isinstance(vintage, (int, float)):
            continue
        data.append({
            'project_name': row[col['Project Name']],
            'project_type': t,
            'registry': registry or 'unknown',
            'vintage_year': int(vintage),
            'credits_issued': float(credits) if credits else 0.0
        })
    return pd.DataFrame(data)

def detect_anomalies(df):
    le_type = LabelEncoder()
    le_registry = LabelEncoder()
    df['type_enc'] = le_type.fit_transform(df['project_type'])
    df['registry_enc'] = le_registry.fit_transform(df['registry'])
    df['credits_log'] = np.log1p(df['credits_issued'])

    features = df[['type_enc', 'registry_enc', 'vintage_year', 'credits_log']]
    model = IsolationForest(contamination=0.05, random_state=42, n_estimators=100)
    df['anomaly_flag'] = model.fit_predict(features)
    df['anomaly_score_raw'] = model.decision_function(features)
    df['is_anomaly'] = df['anomaly_flag'] == -1

    min_s, max_s = df['anomaly_score_raw'].min(), df['anomaly_score_raw'].max()
    df['anomaly_risk_score'] = ((max_s - df['anomaly_score_raw']) / (max_s - min_s) * 100).round(1)
    return df

def main():
    print("Loading Berkeley dataset...")
    df = load_data()
    print(f"Loaded {len(df)} listings")

    print("Running Isolation Forest...")
    df = detect_anomalies(df)

    flagged = df['is_anomaly'].sum()
    print(f"Flagged {flagged} listings ({flagged/len(df)*100:.1f}%) as anomalies")

    output = df[['project_name', 'registry', 'vintage_year', 'is_anomaly', 'anomaly_risk_score']].to_dict('records')
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(output, f, indent=2)

    print(f"Wrote {len(output)} results to {OUTPUT_PATH}")

if __name__ == '__main__':
    main()